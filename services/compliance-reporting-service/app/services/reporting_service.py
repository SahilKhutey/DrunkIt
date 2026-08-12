"""
Compliance reporting service.

Generates regulatory reports:
- Daily transaction summary
- Suspicious activity reports (SAR)
- Audit chain integrity
- Report scheduling
"""

import io
import uuid
from datetime import date, datetime, timezone, timedelta
from typing import Any
import httpx
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from faccp_common.logging import get_logger
from app.config import get_settings
from app.db.models import ComplianceReport, ReportSchedule

logger = get_logger(__name__)
settings = get_settings()


class ComplianceReportingService:
    def __init__(self, db: AsyncSession, http_client: httpx.AsyncClient | None = None) -> None:
        self.db = db
        self._http = http_client or httpx.AsyncClient(timeout=60.0)

    async def generate_daily_transaction_report(
        self, jurisdiction_code: str, report_date: date
    ) -> dict[str, Any]:
        report_id = f"RPT-TXN-{report_date.isoformat()}-{jurisdiction_code}-{uuid.uuid4().hex[:8].upper()}"
        try:
            response = await self._http.get(
                f"{settings.analytics_service_url}/api/v1/metrics/transactions",
                params={
                    "jurisdiction": jurisdiction_code,
                    "from": report_date.isoformat(),
                    "to": (report_date + timedelta(days=1)).isoformat(),
                },
                timeout=30.0,
            )
            data = response.json().get("data", {})
        except Exception as e:
            logger.exception("analytics_fetch_failed", error=str(e))
            data = {}
        workbook = self._build_transaction_workbook(jurisdiction_code, report_date, data)
        excel_bytes = self._workbook_to_bytes(workbook)
        report = ComplianceReport(
            id=str(uuid.uuid4()),
            report_id=report_id,
            report_type="DAILY_TRANSACTION_SUMMARY",
            jurisdiction_code=jurisdiction_code,
            period_start=datetime.combine(report_date, datetime.min.time(), tzinfo=timezone.utc),
            period_end=datetime.combine(report_date + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc),
            generated_at=datetime.now(timezone.utc),
            generated_by="system",
            payload=data,
            file_size_bytes=len(excel_bytes),
            status="ready",
        )
        self.db.add(report)
        await self.db.commit()
        return {
            "report_id": report_id,
            "report_type": "DAILY_TRANSACTION_SUMMARY",
            "jurisdiction": jurisdiction_code,
            "date": report_date.isoformat(),
            "data": data,
            "file_size_bytes": len(excel_bytes),
            "download_url": f"/api/v1/reports/{report_id}/download",
        }

    async def generate_sar_report(
        self, jurisdiction_code: str, period_start: date, period_end: date
    ) -> dict[str, Any]:
        report_id = f"RPT-SAR-{uuid.uuid4().hex[:8].upper()}"
        try:
            response = await self._http.get(
                f"{settings.risk_service_url}/api/v1/risk/high-risk",
                params={"jurisdiction": jurisdiction_code, "from": period_start.isoformat(), "to": period_end.isoformat()},
            )
            risk_data = response.json().get("data", [])
        except Exception:
            risk_data = []
        workbook = self._build_sar_workbook(risk_data)
        excel_bytes = self._workbook_to_bytes(workbook)
        report = ComplianceReport(
            id=str(uuid.uuid4()), report_id=report_id, report_type="SAR",
            jurisdiction_code=jurisdiction_code,
            period_start=datetime.combine(period_start, datetime.min.time(), tzinfo=timezone.utc),
            period_end=datetime.combine(period_end, datetime.min.time(), tzinfo=timezone.utc),
            generated_at=datetime.now(timezone.utc), generated_by="compliance_officer",
            payload={"high_risk_count": len(risk_data), "items": risk_data[:100]},
            file_size_bytes=len(excel_bytes), status="ready",
        )
        self.db.add(report)
        await self.db.commit()
        return {
            "report_id": report_id, "report_type": "SAR",
            "high_risk_count": len(risk_data),
            "file_size_bytes": len(excel_bytes),
        }

    async def generate_audit_chain_integrity_report(self) -> dict[str, Any]:
        try:
            response = await self._http.post(f"{settings.audit_service_url}/api/v1/audit/verify", timeout=120.0)
            data = response.json().get("data", {})
        except Exception as e:
            return {"error": str(e), "verified": False}
        return {
            "report_type": "AUDIT_INTEGRITY",
            "verified_at": datetime.now(timezone.utc).isoformat(),
            "is_intact": data.get("is_intact", False),
            "verified_count": data.get("verified_count", 0),
            "broken_count": data.get("broken_count", 0),
            "broken_events": data.get("broken_events", []),
        }

    async def schedule_report(
        self, report_type: str, jurisdiction_code: str,
        schedule_cron: str, recipients: list[str], generated_by: str
    ) -> ReportSchedule:
        schedule = ReportSchedule(
            id=str(uuid.uuid4()), report_type=report_type,
            jurisdiction_code=jurisdiction_code, schedule_cron=schedule_cron,
            recipients=recipients, is_active=True, created_by=generated_by,
        )
        self.db.add(schedule)
        await self.db.commit()
        return schedule

    def _build_transaction_workbook(self, jurisdiction: str, report_date: date, data: dict) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Summary"
        ws["A1"] = f"Daily Transaction Summary — {jurisdiction} — {report_date.isoformat()}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")
        ws["A3"] = "Metric"
        ws["B3"] = "Value"
        for cell in ws["3:3"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D0D0D0")
        metrics = [
            ("Total Orders", data.get("total_orders", 0)),
            ("Total Revenue (INR)", data.get("total_revenue", 0)),
            ("Unique Consumers", data.get("unique_consumers", 0)),
            ("Unique Retailers", data.get("unique_retailers", 0)),
            ("Age Verification Failures", data.get("age_verification_failures", 0)),
            ("Compliance Denials", data.get("compliance_denials", 0)),
            ("Refunds Issued", data.get("refunds_issued", 0)),
            ("Refund Amount (INR)", data.get("refund_amount", 0)),
        ]
        for i, (metric, value) in enumerate(metrics, start=4):
            ws[f"A{i}"] = metric
            ws[f"B{i}"] = value
        ws["A14"] = "Sales by Category"
        ws["A14"].font = Font(bold=True)
        ws["A15"] = "Category"; ws["B15"] = "Units Sold"; ws["C15"] = "Revenue"
        for c in ws["15:15"]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D0D0D0")
        row = 16
        for cat, stats in (data.get("by_category") or {}).items():
            ws[f"A{row}"] = cat
            ws[f"B{row}"] = stats.get("units", 0)
            ws[f"C{row}"] = stats.get("revenue", 0)
            row += 1
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        return wb

    def _build_sar_workbook(self, risk_items: list[dict]) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Suspicious Activities"
        ws["A1"] = "Suspicious Activity Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")
        headers = ["Date", "Subject ID", "Subject Type", "Risk Level", "Score", "Signals"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D0D0D0")
        for i, item in enumerate(risk_items, start=4):
            ws[f"A{i}"] = item.get("occurred_at", "")
            ws[f"B{i}"] = item.get("subject_id", "")
            ws[f"C{i}"] = item.get("subject_type", "")
            ws[f"D{i}"] = item.get("risk_level", "")
            ws[f"E{i}"] = item.get("risk_score", 0)
            ws[f"F{i}"] = "; ".join(s.get("type", "") for s in item.get("signals", []))
        return wb

    def _workbook_to_bytes(self, wb: Workbook) -> bytes:
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
